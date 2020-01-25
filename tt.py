import openpyxl
import xlrd
import selenium
import time
from selenium import webdriver
import codecs

def nifcheck(numero):
    """ Validação do número de identificação fiscal
    
    >>> valida_nif('999999990') 
    True
    >>> valida_nif('999999999') 
    False
    >>> valida_nif('501442600') # IEFP
    True
    """
    EXPECTED_DIGITS = 9
    if not numero.isdigit() or len(numero) != EXPECTED_DIGITS:
        return False
    soma = sum([int(dig) * (EXPECTED_DIGITS - pos) for pos, dig in enumerate(numero)])
    resto = soma % 11
    if (numero[-1] == '0' and resto == 1):
        resto = (soma + 10) % 11
    if resto == 0:
        return True
    else:
        return False




#init list
Client = []

#open file
wb = xlrd.open_workbook("clcl.xls")

#escolher folha
sh = wb.sheet_by_name(wb.sheet_names()[0])


#extract info
for i in range(2, sh.nrows):
	name = str(sh.cell_value(i,0)).strip()
	nif = str(sh.cell_value(i,3)).strip()
	num = str(sh.cell_value(i,1)).strip()
	
	#remove ".0" from *.0
	if ('.0' in num):
		num = num[0:len(num)-2]

	if ('.0' in nif):
		nif = nif[0:len(nif)-2]

	Client.append({"name" : name, "nif" : nif, "num" : num})
	if i == 648:
		break
	
	
n = 0 
print(Client[5])
for i in Client:
	#if len(i["nif"]) == 9:
	print(nifcheck(i["nif"]))
	
	#verificao de nif 
	if nifcheck(i["nif"]):
		n+=1
	if len(i["nif"])!=0:
		n-=1
	
	#nif invalido é apagado
	if n == -1 :
		i["nif"] = ""
		n += 1
	print(n)


#recheck nif
n = 0 
for i in Client:
	if len(i["nif"])!=0:
		print(i)
		n+=1
		if nifcheck(i["nif"]) == False:
			print("misssssssing")
			break
		
	

print(n)
	

print(len(Client[0]["nif"]))
print(Client[0]["name"])
print(sh.nrows)

count = 0
for c in Client:
	if nifcheck(c["nif"]):
		print(c)
		print(count)
		count+=1

input()

wb = openpyxl.Workbook()

ws = wb._sheets[0]
ws.cell(1,1).value = "Nome"
ws.cell(1,2).value = "Número"
ws.cell(1,3).value = "Nif"

count = 2
for c in Client:
	if not nifcheck(c["nif"]):
		ws.cell(count,1).value = c["name"]
		ws.cell(count,2).value = c["num"]
		count+=1

wb.save("SemNif.xlsx")



#WEB PART--------------------------------------------------
def expand_shadow_element(element):
  shadow_root = driver.execute_script('return arguments[0].shadowRoot', element)
  return shadow_root
  
def wait_secs(sec = 1):
	for i in range(sec):
		time.sleep(1)
		print(i+1)
	





#Login info
def get_email_input ():
	root1 = driver.find_element_by_tag_name('toc-login')
	shadow_root1 = expand_shadow_element(root1)
	
	EmailSenha = shadow_root1.find_element_by_css_selector('casper-login')
	shadow_EmailSenha = expand_shadow_element(EmailSenha)

	#Email text
	Email = shadow_EmailSenha.find_element_by_id('email')
	shadow_Email = expand_shadow_element(Email)

	Email = shadow_Email.find_element_by_id('container')
	shadow_Email = expand_shadow_element(Email)

	Email = Email.find_element_by_id('input-1')
	shadow_Email = expand_shadow_element(Email)


	Email = Email.find_element_by_css_selector('input')
	shadow_Email = expand_shadow_element(Email)
	
	return Email

def get_pass_input ():
	
	root1 = driver.find_element_by_tag_name('toc-login')
	shadow_root1 = expand_shadow_element(root1)
	
	EmailSenha = shadow_root1.find_element_by_css_selector('casper-login')
	shadow_EmailSenha = expand_shadow_element(EmailSenha)
	
	#password text
	Pass = shadow_EmailSenha.find_element_by_id('password')
	shadow_Email = expand_shadow_element(Pass)

	Pass = shadow_Email.find_element_by_id('container')
	shadow_Email = expand_shadow_element(Pass)

	Pass = Pass.find_element_by_id('input-2')
	shadow_Email = expand_shadow_element(Pass)

	Pass = Pass.find_element_by_css_selector('input')
	shadow_Email = expand_shadow_element(Pass)
	
	return Pass

def get_login_button ():
	root1 = driver.find_element_by_tag_name('toc-login')
	shadow_root1 = expand_shadow_element(root1)
	
	EmailSenha = shadow_root1.find_element_by_css_selector('casper-login')
	shadow_EmailSenha = expand_shadow_element(EmailSenha)
	
	return shadow_EmailSenha.find_element_by_css_selector('casper-button')
#/Login info	


#LogIn
def toc_login():
	driver.get("https://app2.toconline.pt/login")
	wait_secs(3)
	
	email = input("Email:")
	senha = input("Senha:")
	get_email_input().send_keys(email)
	get_pass_input().send_keys(senha)
	
	get_login_button().click()
	
	while driver.current_url != 'https://app2.toconline.pt/toc/my_company/summary':
		 wait_secs(3)
	
def get_add_client_doc():
	page = driver.find_element_by_css_selector('toc-app')
	shadow_page = expand_shadow_element(page)
	
	page = shadow_page.find_element_by_css_selector('toc-rorframe')
	shadow_page = expand_shadow_element(page)

	print(shadow_page)
	
	iframe = shadow_page.find_elements_by_tag_name('iframe')[0]
	
	driver.switch_to_frame(iframe)

def click_add_client_button():

	bt = driver.find_element_by_xpath('//*[@id="new_customer"]/i')
	driver.execute_script("arguments[0].click()", bt)
	
	wait_secs(2)
	
def fill_add_client_inputs(nif, name, num):
	
	nif_input 		= driver.find_element_by_xpath('//*[@id="customer_tax_registration_number"]')	
	number_input 	= driver.find_element_by_xpath('//*[@id="customer_contact_name"]')
	name_input 		= driver.find_element_by_xpath('//*[@id="customer_business_name"]')
	
	nif_input.send_keys(nif)
	name_input.send_keys(name)
	number_input.send_keys(num) 
	
def click_save_client_button():
	
	bt = driver.find_element_by_xpath('//*[@id="new_customer"]/div[4]/div/div[1]/input')
	driver.execute_script("arguments[0].click()", bt)

	wait_secs(2)
	try:
		bt = driver.find_element_by_xpath('//*[@id="modal_form"]/div[2]/div/div/form/div[2]/div/a[1]')
		driver.execute_script("arguments[0].click()", bt)
		wait_secs(2)
	except:
		pass
		
	
def add_client(nif, name, num):
	
	click_add_client_button()
	fill_add_client_inputs(nif, name, num)
	click_save_client_button()

	


#Open Browser
driver = webdriver.Chrome()
wait_secs(1)

#Login in Toc
toc_login()

#open Costumers page
driver.get("https://app2.toconline.pt/toc/my_company/customers")
wait_secs(6)

#Get the "add_costumer" file
get_add_client_doc()

#add client
count = 0

for c in Client:
	if nifcheck(c["nif"]):
		count+=1
		add_client(c["nif"], c["name"], c["num"])
		print(count)

